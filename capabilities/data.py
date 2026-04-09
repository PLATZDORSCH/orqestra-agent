"""Data file reader — read CSV, Excel, JSON, and text files.

Provides a simple way for the agent to inspect structured data without
writing a full Python script. Returns the data as JSON with preview rows,
column info, and basic statistics for numeric columns.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from pathlib import Path
from typing import Any

from core.capabilities import Capability

log = logging.getLogger(__name__)

_MAX_PREVIEW_ROWS = 50
_MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


def _read_csv(path: Path, delimiter: str | None = None) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(text[:4096])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = []
    for i, row in enumerate(reader):
        if i >= _MAX_PREVIEW_ROWS:
            break
        rows.append(dict(row))

    total = sum(1 for _ in csv.reader(io.StringIO(text))) - 1
    columns = reader.fieldnames or []
    stats = _compute_stats(rows, columns)

    return {
        "format": "csv",
        "columns": columns,
        "total_rows": total,
        "preview_rows": len(rows),
        "data": rows,
        "stats": stats,
    }


def _read_json(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))

    if isinstance(data, list):
        preview = data[:_MAX_PREVIEW_ROWS]
        columns = sorted({k for row in preview if isinstance(row, dict) for k in row})
        return {
            "format": "json",
            "type": "array",
            "total_items": len(data),
            "preview_rows": len(preview),
            "columns": columns,
            "data": preview,
            "stats": _compute_stats(preview, columns) if columns else {},
        }
    elif isinstance(data, dict):
        return {
            "format": "json",
            "type": "object",
            "keys": list(data.keys()),
            "data": data,
        }
    else:
        return {"format": "json", "type": type(data).__name__, "data": data}


def _read_excel(path: Path, sheet: str | None = None) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet] if sheet and sheet in sheet_names else wb.active

    rows: list[dict[str, Any]] = []
    headers: list[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
            continue
        if i > _MAX_PREVIEW_ROWS:
            break
        rows.append({h: v for h, v in zip(headers, row)})

    total = ws.max_row - 1 if ws.max_row else 0
    stats = _compute_stats(rows, headers)

    wb.close()
    return {
        "format": "excel",
        "sheets": sheet_names,
        "active_sheet": ws.title,
        "columns": headers,
        "total_rows": total,
        "preview_rows": len(rows),
        "data": rows,
        "stats": stats,
    }


def _read_text(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    return {
        "format": "text",
        "total_lines": len(lines),
        "preview_lines": min(len(lines), 100),
        "content": "\n".join(lines[:100]),
    }


def _compute_stats(rows: list[dict], columns: list[str]) -> dict[str, Any]:
    """Compute basic stats for numeric columns."""
    stats: dict[str, Any] = {}
    for col in columns:
        values = []
        for row in rows:
            v = row.get(col)
            if v is None:
                continue
            try:
                values.append(float(v))
            except (ValueError, TypeError):
                continue
        if len(values) >= 2:
            stats[col] = {
                "count": len(values),
                "min": round(min(values), 2),
                "max": round(max(values), 2),
                "mean": round(sum(values) / len(values), 2),
            }
    return stats


def _handle_read_data(args: dict) -> str:
    path = Path(args["path"]).expanduser().resolve()

    if not path.exists():
        return json.dumps({"error": f"File not found: {args['path']}"})
    if not path.is_file():
        return json.dumps({"error": f"Not a file: {args['path']}"})
    if path.stat().st_size > _MAX_FILE_SIZE:
        return json.dumps({"error": f"File too large ({path.stat().st_size:,} bytes, max {_MAX_FILE_SIZE:,})"})

    suffix = path.suffix.lower()

    try:
        if suffix == ".csv":
            result = _read_csv(path, delimiter=args.get("delimiter"))
        elif suffix in (".xlsx", ".xls"):
            result = _read_excel(path, sheet=args.get("sheet"))
        elif suffix == ".json":
            result = _read_json(path)
        elif suffix in (".tsv",):
            result = _read_csv(path, delimiter="\t")
        else:
            result = _read_text(path)
    except Exception as exc:
        return json.dumps({"error": f"{type(exc).__name__}: {exc}"})

    return json.dumps(result, ensure_ascii=False, default=str)


read_data = Capability(
    name="read_data",
    description=(
        "Read a data file and return its contents with column info and basic statistics. "
        "Supports CSV, TSV, Excel (.xlsx), JSON, and plain text files. "
        "For large files, returns the first 50 rows as preview plus total row count."
    ),
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to the data file (absolute or relative to working directory)"},
            "delimiter": {"type": "string", "description": "CSV delimiter (auto-detected if not specified)"},
            "sheet": {"type": "string", "description": "Excel sheet name (uses active sheet if not specified)"},
        },
        "required": ["path"],
    },
    handler=_handle_read_data,
)
